import io
from datetime import datetime, timedelta

from dateutil.relativedelta import relativedelta
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.services import (
    emotional_release_service,
    energy_knot_service,
    group_case_service,
    internal_course_service,
    membership_card_service,
    offline_course_service,
    oh_card_reading_service,
    organization_service,
    other_project_service,
    tea_seat_fee_service,
)

PROJECT_SOURCES = (
    ("membership_card", "会员卡", membership_card_service.list_cards),
    ("group_case", "觉醒游戏", group_case_service.list_cases),
    ("emotional_release", "情绪释放", emotional_release_service.list_releases),
    ("oh_card_reading", "OH卡诊断", oh_card_reading_service.list_readings),
    ("energy_knot", "能量结", energy_knot_service.list_knots),
    ("internal_course", "内部课程", internal_course_service.list_courses),
    ("tea_seat_fee", "茶位费", tea_seat_fee_service.list_fees),
    ("offline_course", "线下落地课程", offline_course_service.list_courses),
    ("other", "其他项目", other_project_service.list_projects),
)

HEADERS = [
    "成交日期",
    "付费类型",
    "用户昵称",
    "项目内容",
    "数量/期限",
    "金额",
    "生效日期",
    "到期日期",
    "成交人",
    "支付方式",
    "所属组织",
    "备注",
    "创建人",
    "录入时间",
]

COLUMN_WIDTHS = [13, 16, 16, 30, 14, 14, 13, 13, 24, 14, 18, 30, 14, 20]


def _offline_expiry(item: dict) -> str:
    effective_date = item.get("effective_date") or ""
    validity_value = int(item.get("validity_value") or 0)
    if not effective_date or validity_value <= 0:
        return ""
    try:
        start = datetime.strptime(effective_date[:10], "%Y-%m-%d")
    except ValueError:
        return ""
    if item.get("validity_unit") == "day":
        end = start + timedelta(days=validity_value) - timedelta(days=1)
    else:
        end = start + relativedelta(months=validity_value) - timedelta(days=1)
    return end.strftime("%Y-%m-%d")


def _project_detail(item: dict, project_type: str, type_label: str) -> str:
    if project_type == "membership_card":
        return item.get("card_type") or ""
    if project_type == "internal_course":
        return item.get("course_type") or ""
    if project_type == "oh_card_reading":
        teacher = item.get("diagnosis_teacher") or ""
        return f"诊断老师：{teacher}" if teacher else type_label
    if project_type == "other":
        return " / ".join(filter(None, [item.get("category"), item.get("project_name")]))
    return type_label


def _project_quantity(item: dict, project_type: str) -> str:
    if project_type == "membership_card":
        total_count = item.get("total_count")
        return "不限" if total_count is None else f"{total_count}次"
    if project_type in {"group_case", "emotional_release"}:
        return f"{item.get('purchase_count') or 0}次"
    if project_type == "energy_knot":
        return f"{item.get('purchase_count') or 0}个"
    if project_type == "oh_card_reading":
        return f"{(item.get('diagnosis_duration') or 1) * 0.5:g}小时"
    if project_type == "tea_seat_fee":
        return f"{item.get('quantity') or 1}位"
    if project_type == "offline_course":
        unit = "天" if item.get("validity_unit") == "day" else "个月"
        return f"{item.get('validity_value') or 1}{unit}"
    if project_type == "other":
        total_count = item.get("total_count")
        if total_count is None:
            total_count = item.get("remaining_count")
        return "不限" if total_count is None else f"{total_count}次"
    return ""


def _amount(item: dict, project_type: str) -> float:
    if project_type in {"membership_card", "internal_course"}:
        return float(item.get("price") or 0)
    if project_type == "other":
        return float(item.get("fee") or 0)
    return float(item.get("amount") or 0)


def _closer_names(item: dict) -> str:
    closers = item.get("closers") or []
    if closers:
        return "、".join(
            f"{closer.get('name') or ''} ¥{float(closer.get('amount') or 0):g}"
            for closer in closers
        )
    return item.get("closer_name") or ""


def _format_created_at(value) -> str:
    if not value:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    return str(value).replace("T", " ")[:19]


def _build_rows(
    user_role: str,
    date_from: str | None = None,
    date_to: str | None = None,
) -> list[dict]:
    organization_names = {
        organization.id: organization.name
        for organization in organization_service.list_organizations()
    }
    rows = []
    for project_type, type_label, list_records in PROJECT_SOURCES:
        for record in list_records():
            item = record.model_dump(mode="json")
            created_at = item.get("created_at") or ""
            deal_date = item.get("deal_date") or ""
            normalized_deal_date = str(deal_date)[:10]
            if date_from and (not normalized_deal_date or normalized_deal_date < date_from):
                continue
            if date_to and (not normalized_deal_date or normalized_deal_date > date_to):
                continue
            effective_date = item.get("effective_date") or ""
            expiry_date = (
                _offline_expiry(item)
                if project_type == "offline_course"
                else (item.get("expiry_date") or "")
            )
            rows.append({
                "成交日期": deal_date,
                "付费类型": type_label,
                "用户昵称": item.get("nickname") or "",
                "项目内容": _project_detail(item, project_type, type_label),
                "数量/期限": _project_quantity(item, project_type),
                "金额": _amount(item, project_type),
                "生效日期": effective_date,
                "到期日期": expiry_date,
                "成交人": _closer_names(item),
                "支付方式": item.get("payment_method") or "",
                "所属组织": organization_names.get(item.get("organization_id"), ""),
                "备注": item.get("notes") or "",
                "创建人": item.get("created_by") or "",
                "录入时间": _format_created_at(created_at),
                "_sort_date": deal_date or str(created_at)[:10],
                "_sort_created_at": str(created_at),
            })
    rows.sort(
        key=lambda row: (row["_sort_date"], row["_sort_created_at"]),
        reverse=True,
    )
    return rows


def build_payment_export(
    user_role: str,
    date_from: str | None = None,
    date_to: str | None = None,
) -> tuple[io.BytesIO, int]:
    rows = _build_rows(user_role, date_from, date_to)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "全部付费记录"
    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "A2"

    for index, width in enumerate(COLUMN_WIDTHS, 1):
        worksheet.column_dimensions[get_column_letter(index)].width = width

    header_fill = PatternFill(fill_type="solid", fgColor="F7F8FA")
    thin_side = Side(style="thin", color="E8E8E8")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    for column, header in enumerate(HEADERS, 1):
        cell = worksheet.cell(row=1, column=column, value=header)
        cell.font = Font(bold=True, color="2B2F36")
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center")
    worksheet.row_dimensions[1].height = 24

    for row_index, row_data in enumerate(rows, 2):
        for column, header in enumerate(HEADERS, 1):
            cell = worksheet.cell(row=row_index, column=column, value=row_data[header])
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
            if header == "金额":
                cell.number_format = '¥#,##0.00'

    worksheet.auto_filter.ref = f"A1:N{max(1, len(rows) + 1)}"
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output, len(rows)
