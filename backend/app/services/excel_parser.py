import json
import openpyxl
from io import BytesIO
from langchain_openai import ChatOpenAI
from langchain_core.messages import HumanMessage, SystemMessage

from app.models.customer import CustomerCreate
from app.services.customer_ai_config_service import get_config as get_customer_ai_config
from app.config.settings import settings

PARSE_PROMPT = """你是一个客户信息提取助手。用户会给你一个表格的内容（以 JSON 格式表示），你需要从每一行中提取客户信息，以 JSON 数组格式返回。

每行的格式如下：
{
  "nickname": "客户昵称",
  "name": "客户姓名",
  "phone": "电话号码",
  "wechat": "微信号",
  "age": "年龄",
  "referrer": "引流人/推荐人",
  "paid_content": [
    {"type": "399次卡", "usage_count": 0, "salesperson": "成交人"}
  ],
  "visit_count": 0,
  "core_situation": "核心情况摘要",
  "need_tags": "需求标签",
  "follow_up_node": "跟进节点",
  "follow_up_action": "核心动作",
  "positions": ["课程部"],
  "self_tags": ["自我成长"]
}

paid_content 的 type 只能是：399次卡、3999会员、半年卡、2w疗愈师
positions 的值只能是：课程部、流量部、承接部、售后部、成就君、信息管理（可多选）
self_tags 的值只能是：自我成长、共创、变现

无法从表格中提取的字段留空字符串或空数组。只返回 JSON 数组，不要其他内容。"""


def _sanitize_cell(value: str) -> str:
    """防止 CSV/公式注入：转义公式前缀字符，清理危险字符"""
    if not value:
        return value
    stripped = value.lstrip()
    if stripped and stripped[0] in ("=", "+", "-", "@", "\t", "\r"):
        return "'" + value
    # 清理可能破坏 CSV/markdown 解析的字符
    value = value.replace("|", "\\|").replace("\n", " ").replace("\r", "")
    return value


def parse_excel(file_content: bytes) -> list[CustomerCreate]:
    config = get_customer_ai_config()
    api_key = config.api_key or settings.llm_api_key
    base_url = config.base_url or settings.llm_base_url
    model = config.model or settings.llm_model
    system_prompt = config.system_prompt or PARSE_PROMPT

    # 读取 Excel 文件
    wb = openpyxl.load_workbook(BytesIO(file_content), read_only=True)
    ws = wb.active

    # 获取表头
    headers = []
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        headers.append(str(cell.value) if cell.value else "")

    # 获取数据行
    MAX_ROWS = 200
    MAX_CHARS = 80000
    rows = []
    total_chars = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_data = {}
        for i, value in enumerate(row):
            if i < len(headers) and headers[i]:
                cell_str = _sanitize_cell(str(value)) if value is not None else ""
                row_data[headers[i]] = cell_str
                total_chars += len(cell_str)
        if any(row_data.values()):  # 跳过空行
            rows.append(row_data)
            if len(rows) > MAX_ROWS:
                wb.close()
                raise ValueError(f"数据行数超过上限（最多 {MAX_ROWS} 行），请精简后重新上传")
            if total_chars > MAX_CHARS:
                wb.close()
                raise ValueError(f"数据量过大（最多 {MAX_CHARS // 1000}K 字符），请精简后重新上传")

    wb.close()

    if not rows:
        return []

    # 调用 AI 解析
    llm = ChatOpenAI(
        model=model,
        api_key=api_key,
        base_url=base_url,
        temperature=0,
        max_tokens=4096,
    )

    response = llm.invoke([
        SystemMessage(content=system_prompt),
        HumanMessage(content=f"请从以下表格数据中提取客户信息：\n\n{json.dumps(rows, ensure_ascii=False, indent=2)}"),
    ])

    content = response.content

    if not content or not content.strip():
        raise ValueError("AI 返回了空内容，请检查 API Key 和模型配置")

    # 清理内容
    content = content.strip()
    if content.startswith("```"):
        content = content.split("\n", 1)[1].rsplit("```", 1)[0]
    if content.startswith("json"):
        content = content[4:]
    content = content.strip()

    try:
        data = json.loads(content)
    except json.JSONDecodeError:
        raise ValueError("AI 返回的内容不是有效的 JSON 格式")

    if not isinstance(data, list):
        raise ValueError(f"AI 返回的内容不是数组格式: {type(data)}")

    return [CustomerCreate(**item) for item in data]
