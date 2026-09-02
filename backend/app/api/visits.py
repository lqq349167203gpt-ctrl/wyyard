import io

from fastapi import APIRouter, HTTPException, Query, Request
from fastapi.responses import StreamingResponse

from app.models.visit import VisitRecordCreate
from app.services import customer_access_service, customer_service, visit_service, visit_verification_service
from app.services.customer_service import get_customer
from app.utils.pagination import paginate
from app.utils.record_ownership import ensure_creator_for_changed_fields, ensure_record_creator, stamp_creator
from app.utils.request_roles import get_request_roles

router = APIRouter(prefix="/api/visits", tags=["visits"])

VISIT_NOTE_CATEGORY_LABELS = {
    "visit_need": "来访需求",
    "customer_info": "客户信息",
    "follow_up": "跟进点",
}


def _visible_customer_ids(request: Request | None) -> set[str] | None:
    if request is None:
        return None
    return customer_access_service.visible_customer_ids(
        request,
        customer_service.list_customers(),
    )


def _filter_visible_visits(request: Request | None, records: list) -> list:
    visible_ids = _visible_customer_ids(request)
    if visible_ids is None:
        return records
    return [record for record in records if record.customer_id in visible_ids]


def _fill_member_type(record, private_need: str = ""):
    """从客户信息实时填充会员身份、昵称和引流来源"""
    data = record.model_dump(mode="json")
    customer = get_customer(record.customer_id)
    data["member_type"] = customer.member_type if customer else ""
    data["nickname"] = customer.nickname if customer else ""
    data["referrer"] = customer.referrer if customer else ""
    data["needs"] = private_need
    return data


def _private_need_map(request: Request | None, visit_ids: list[str]) -> dict[str, str]:
    if request is None or not visit_ids:
        return {}
    from app.services import visit_note_service

    notes = visit_note_service.list_visible_notes(
        visit_ids,
        getattr(request.state, "user_id", "") or "",
        getattr(request.state, "user_owner", "") or "",
        getattr(request.state, "user_name", "") or "",
    )
    result: dict[str, list[str]] = {}
    for note in notes:
        if note.category != "visit_need":
            continue
        creator = note.created_by or "历史记录"
        result.setdefault(note.visit_id, []).append(f"{creator}：{note.content}")
    return {visit_id: "\n".join(lines) for visit_id, lines in result.items()}


def _visit_note_maps(request: Request | None, visit_ids: list[str]) -> tuple[dict[str, str], dict[str, list[dict]]]:
    """一次读取可见信息，同时生成兼容需求文本和表格所需的信息列表。"""
    if request is None or not visit_ids:
        return {}, {}
    from app.services import visit_note_service

    account_id = getattr(request.state, "user_id", "") or ""
    owner_name = getattr(request.state, "user_owner", "") or ""
    username = getattr(request.state, "user_name", "") or ""
    notes = visit_note_service.list_visible_notes(
        visit_ids,
        account_id,
        owner_name,
        username,
    )
    need_lines: dict[str, list[str]] = {}
    note_map: dict[str, list[dict]] = {}
    for note in notes:
        can_manage = visit_note_service.can_manage_note(
            note,
            account_id,
            owner_name,
            username,
        )
        item = note.model_dump(mode="json")
        item["category_label"] = VISIT_NOTE_CATEGORY_LABELS[note.category]
        item["can_edit"] = can_manage
        item["can_delete"] = can_manage
        note_map.setdefault(note.visit_id, []).append(item)
        if note.category == "visit_need":
            creator = note.created_by or "历史记录"
            need_lines.setdefault(note.visit_id, []).append(f"{creator}：{note.content}")
    return (
        {visit_id: "\n".join(lines) for visit_id, lines in need_lines.items()},
        note_map,
    )


def _fill_daily_amount(items: list, date: str):
    """汇总每个客户当日成交金额，注入 daily_amount 字段"""
    if not date:
        return
    from app.services import (
        emotional_release_service,
        energy_knot_service,
        group_case_service,
        internal_course_service,
        membership_card_service,
        oh_card_reading_service,
        other_project_service,
    )

    def _get_amount(r):
        return getattr(r, "fee", None) or getattr(r, "price", None) or getattr(r, "amount", None) or 0

    def _get_date(r):
        d = getattr(r, "deal_date", None)
        if d:
            return d
        ca = getattr(r, "created_at", None)
        if ca:
            if hasattr(ca, "strftime"):
                return ca.strftime("%Y-%m-%d")
            return str(ca)[:10]
        return ""

    # 按 customer_id 汇总当日金额
    amount_map: dict[str, float] = {}
    services = [
        membership_card_service.list_cards(),
        group_case_service.list_cases(),
        emotional_release_service.list_releases(),
        energy_knot_service.list_knots(),
        internal_course_service.list_courses(),
        oh_card_reading_service.list_readings(),
        other_project_service.list_projects(),
    ]
    for records in services:
        for r in records:
            if getattr(r, "voided", False):
                continue
            if getattr(r, "is_deleted", False):
                continue
            if _get_date(r) == date:
                cid = getattr(r, "customer_id", "")
                if cid:
                    amount_map[cid] = amount_map.get(cid, 0) + _get_amount(r)

    for item in items:
        item["daily_amount"] = amount_map.get(item.get("customer_id", ""), 0)


@router.get("")
async def list_visits(
    date: str = None,
    customer_id: str = None,
    space_id: str = None,
    page: int | None = Query(None, ge=1),
    page_size: int | None = Query(None, ge=1, le=100),
    request: Request = None,
):
    records = _filter_visible_visits(
        request,
        visit_service.list_visits(date, customer_id, space_id),
    )
    need_map, note_map = _visit_note_maps(request, [record.id for record in records])
    items = []
    for record in records:
        item = _fill_member_type(record, need_map.get(record.id, ""))
        item["visit_notes"] = note_map.get(record.id, [])
        items.append(item)
    role = get_request_roles(request) if request else ["超级管理员"]
    if date and customer_access_service.can_view_transaction_summary(role):
        _fill_daily_amount(items, date)
    if page is not None:
        return paginate(items, page, page_size or 10)
    return items


@router.get("/light")
async def list_visits_light(
    date: str = None,
    space_id: str = None,
    request: Request = None,
):
    """轻量版列表，不计算活动详情，适用于主页快速加载"""
    try:
        items = visit_service.list_visits_light(date, space_id)
        need_map = _private_need_map(request, [str(item.get("id") or "") for item in items])
        for item in items:
            item["needs"] = need_map.get(str(item.get("id") or ""), "")
        visible_ids = _visible_customer_ids(request)
        if visible_ids is None:
            return items
        return [item for item in items if item.get("customer_id") in visible_ids]
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/counts")
async def get_visit_counts(
    customer_ids: str | None = Query(None),
    start_date: str | None = Query(None),
    end_date: str | None = Query(None),
    member_types: str | None = Query(None),
    space_id: str | None = Query(None),
    request: Request = None,
):
    """返回各日期的到场人数统计 {date: count}，不做活动计数。支持 member_types 过滤权限"""
    ids = None
    if member_types is not None:
        types = [m for m in member_types.split(",") if m]
        if types:
            customers = customer_service.list_all_customers()
            ids = [c.id for c in customers if c.member_type in types]
            if not ids:
                return {}
        else:
            ids = None  # member_types 参数为空 → 超管，不过滤
    elif customer_ids is not None:
        ids = [x for x in customer_ids.split(",") if x]
        if not ids:
            return {}
    visible_ids = _visible_customer_ids(request)
    if visible_ids is not None:
        ids = list(visible_ids if ids is None else visible_ids.intersection(ids))
        if not ids:
            return {}
    return visit_service.get_date_counts(ids, start_date, end_date, space_id)


@router.get("/search-customers")
async def search_customers(q: str = "", request: Request = None):
    results = visit_service.search_customers(q)
    visible_ids = _visible_customer_ids(request)
    if visible_ids is None:
        return results
    return [result for result in results if result.id in visible_ids]


@router.post("/reorder")
async def reorder_visits(data: dict, request: Request):
    """批量更新排序权重。传入 {"ids": ["id1", "id2", ...]} 按顺序设置 sort_order"""
    ids = data.get("ids", [])
    if not ids:
        raise HTTPException(status_code=400, detail="ids 不能为空")
    records = {record.id: record for record in visit_service.list_visits()}
    if any(visit_id not in records for visit_id in ids):
        raise HTTPException(status_code=404, detail="邀约记录不存在")
    for visit_id in ids:
        visit_verification_service.ensure_record_unverified(records[visit_id])
    visible_ids = _visible_customer_ids(request)
    if visible_ids is not None and any(
        records[visit_id].customer_id not in visible_ids for visit_id in ids
    ):
        raise HTTPException(status_code=403, detail="只能调整可见客户的邀约顺序")
    visit_service.reorder_visits(ids)
    return {"message": "排序已更新"}


@router.get("/export")
async def export_visits(date: str = None, space_id: str = None, request: Request = None):
    """导出邀约到场 xlsx，格式与 PC 端一致"""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    records = _filter_visible_visits(
        request,
        visit_service.list_visits(date, space_id=space_id),
    )
    active_records = [record for record in records if not record.cancelled]
    need_map = _private_need_map(request, [record.id for record in active_records])
    items = [_fill_member_type(record, need_map.get(record.id, "")) for record in active_records]

    # 构建组长映射（复用 PC 端逻辑）
    role_map = {}
    for r in items:
        if r.get("is_leader"):
            role_map[r["customer_id"]] = "组长"

    rows = []
    for v in items:
        role = role_map.get(v["customer_id"], "")
        rows.append({
            "引流人": v.get("referrer") or "-",
            "客户昵称": v.get("nickname") or "",
            "预计时间": v.get("visit_time") or "",
            "参与次数": v.get("visit_count") or 0,
            "会员身份": v.get("member_type") or "",
            "来访需求": v.get("needs") or "",
            "组长情况": role or "-",
            "组长获得的信息": "",
            "邀约人": v.get("referrer_handler") or "",
        })

    if not rows:
        # 空表也返回有效 xlsx（只有表头）
        rows = []

    headers = ["引流人", "客户昵称", "预计时间", "参与次数", "会员身份", "来访需求", "组长情况", "组长获得的信息", "邀约人"]
    col_widths = [10, 12, 10, 10, 12, 40, 10, 30, 10]

    wb = Workbook()
    ws = wb.active
    ws.title = "邀约到场"
    ws.sheet_view.showGridLines = False

    # 设置列宽
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    # 写表头
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D0D3D6", end_color="D0D3D6", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="C0C4CC"),
        right=Side(style="thin", color="C0C4CC"),
        top=Side(style="thin", color="C0C4CC"),
        bottom=Side(style="thin", color="C0C4CC"),
    )

    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    # 写数据行
    for r, row_data in enumerate(rows, 2):
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=r, column=c, value=row_data[h])
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    # 输出到内存
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    # HTTP 头部只支持 latin-1，中文文件名需按 RFC 5987 用 filename* 百分号编码
    from urllib.parse import quote
    if date:
        y, m, d = date.split('-')
        date_text = f"{int(y)}年{int(m)}月{int(d)}日"
    else:
        date_text = "全部"
    filename = f"{date_text}邀约名单.xlsx"
    disposition = f"attachment; filename=\"visits_{date or 'all'}.xlsx\"; filename*=UTF-8''{quote(filename)}"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": disposition},
    )


@router.get("/{visit_id}")
async def get_visit(visit_id: str, request: Request):
    record = visit_service.get_visit(visit_id)
    if not record:
        raise HTTPException(status_code=404, detail="记录不存在")
    customer_access_service.require_customer_scope(request, record.customer_id)
    need_map, note_map = _visit_note_maps(request, [record.id])
    item = _fill_member_type(record, need_map.get(record.id, ""))
    item["visit_notes"] = note_map.get(record.id, [])
    return item


@router.post("")
async def create_visit(data: VisitRecordCreate, request: Request):
    customer_access_service.require_customer_scope(request, data.customer_id, action="邀约")
    visit_verification_service.ensure_scope_unverified(data.visit_date, data.space_id or "")
    try:
        initial_need = str(data.needs or "").strip()
        create_data = data.model_copy(update={"needs": ""}) if initial_need else data
        record = visit_service.create_visit(stamp_creator(create_data, request))
        if initial_need:
            from app.services import visit_note_service

            visit_note_service.create_note(
                record.id,
                "visit_need",
                initial_need,
                creator_id=getattr(request.state, "user_id", "") or "",
                creator=(
                    getattr(request.state, "user_owner", "")
                    or getattr(request.state, "user_name", "")
                    or ""
                ),
            )
        return _fill_member_type(record, initial_need)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.patch("/{visit_id}")
async def update_visit(visit_id: str, data: dict, request: Request):
    # 记录更新前的到场状态
    old_record = visit_service.get_visit(visit_id)
    if not old_record:
        raise HTTPException(status_code=404, detail="记录不存在")
    customer_access_service.require_customer_scope(request, old_record.customer_id, action="修改")
    if data.get("customer_id") and data["customer_id"] != old_record.customer_id:
        customer_access_service.require_customer_scope(request, data["customer_id"], action="邀约")
    visit_verification_service.ensure_update_allowed(old_record, data)
    if "needs" in data and (old_record.cancelled or "cancelled" in data):
        raise HTTPException(status_code=400, detail="取消或恢复邀约时不能同时修改来访需求")
    ensure_creator_for_changed_fields(
        request,
        old_record,
        data,
        {
            "visit_date",
            "visit_time",
            "customer_id",
            "referrer_handler",
        },
        "邀约的客户、邀约人或时间",
        "visits",
    )
    old_arrived = old_record.arrived if old_record else False

    # 兼容旧客户端：客户信息/跟进点不再覆盖整段文本，而是按当前账号追加为独立记录。
    collaboration_fields = {
        "needs": "visit_need",
        "feedback": "customer_info",
        "healing_notes": "follow_up",
    }
    collaboration_updates = []
    for field, category in collaboration_fields.items():
        if field not in data:
            continue
        content = str(data.pop(field) or "").strip()
        current = str(getattr(old_record, field, "") or "").strip() if old_record else ""
        if content and content != current:
            collaboration_updates.append((category, content))

    try:
        record = visit_service.update_visit(visit_id, data) if data else old_record
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    if not record:
        raise HTTPException(status_code=404, detail="记录不存在")

    if collaboration_updates:
        from app.services import visit_note_service

        creator_id = getattr(request.state, "user_id", "") or ""
        creator = (
            getattr(request.state, "user_owner", "")
            or getattr(request.state, "user_name", "")
            or ""
        )
        for category, content in collaboration_updates:
            visit_note_service.create_note(
                visit_id,
                category,
                content,
                creator_id=creator_id,
                creator=creator,
            )
        record = visit_service.get_visit(visit_id)

    # 从未到场 → 已到场：发送通知给客户
    if not old_arrived and record.arrived:
        from app.services import client_notification_service, customer_service
        customer = customer_service.get_customer(record.customer_id)
        if customer:
            activity_names = "、".join(a.name for a in (record.activities or []) if a.name) or "今日活动"
            client_notification_service.create_notification(
                customer_id=record.customer_id,
                type="arrival_confirmed",
                title="已确认到场",
                content=f'您在{record.visit_date}的{activity_names}已确认到场',
                activity_name=activity_names,
                activity_date=record.visit_date,
                operator="管理员",
            )

    need_map = _private_need_map(request, [record.id])
    return _fill_member_type(record, need_map.get(record.id, ""))


@router.delete("/{visit_id}")
async def delete_visit(visit_id: str, request: Request):
    record = visit_service.get_visit(visit_id)
    if not record:
        raise HTTPException(status_code=404, detail="记录不存在")
    customer_access_service.require_customer_scope(request, record.customer_id, action="删除")
    visit_verification_service.ensure_record_unverified(record)
    ensure_record_creator(request, record, "邀约", "visits")
    if not visit_service.delete_visit(visit_id):
        raise HTTPException(status_code=404, detail="记录不存在")
    return {"message": "已删除"}
