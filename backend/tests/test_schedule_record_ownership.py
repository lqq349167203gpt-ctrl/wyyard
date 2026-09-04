"""课表与邀约记录按字段校验创建人权限。"""

import uuid
from io import BytesIO

from openpyxl import load_workbook

from app.services import position_edit_permission_service


def _create_other_account_headers(client):
    suffix = uuid.uuid4().hex[:10]
    password = f"pw{suffix}9"
    position_response = client.post("/api/positions", json={
        "name": f"归属测试角色_{suffix}",
    })
    assert position_response.status_code == 200
    position = position_response.json()
    permission_response = client.put("/api/position-permissions/full", json={
        "position": position["name"],
        "pages": ["class-records", "daily-activities"],
        "edit_permissions": {"visits": "own", "activities": "own"},
    })
    assert permission_response.status_code == 200
    response = client.post("/api/accounts", json={
        "owner": f"其他员工_{suffix}",
        "role": position["name"],
        "username": f"ownership_{suffix}",
        "password": password,
        "enabled": True,
    })
    assert response.status_code == 200
    account = response.json()
    login = client.post("/api/accounts/login", json={
        "username": account["username"],
        "password": password,
    })
    assert login.status_code == 200
    return account, position, {"Authorization": f"Bearer {login.json()['token']}"}


def test_visit_receptionist_and_goal_are_shared_for_non_view_accounts(client, created_customer):
    visit_response = client.post("/api/visits", json={
        "visit_date": "2026-09-04",
        "visit_time": "14:00",
        "customer_id": created_customer["id"],
        "receptionist": created_customer["nickname"],
        "goal": "完成首次需求确认",
    })
    assert visit_response.status_code == 200, visit_response.text
    visit = visit_response.json()
    assert visit["receptionist"] == created_customer["nickname"]
    assert visit["goal"] == "完成首次需求确认"

    account, position, other_headers = _create_other_account_headers(client)
    try:
        shared_update = client.patch(
            f"/api/visits/{visit['id']}",
            json={"receptionist": "", "goal": "由接待人完成方案沟通"},
            headers=other_headers,
        )
        assert shared_update.status_code == 200, shared_update.text
        assert shared_update.json()["receptionist"] == ""
        assert shared_update.json()["goal"] == "由接待人完成方案沟通"

        visit_logs = client.get("/api/operation-logs", params={"entity_id": visit["id"]})
        assert visit_logs.status_code == 200
        assert any(
            "目标" in item["content"] and "接待人" in item["content"]
            for item in visit_logs.json()
        )

        invalid_receptionist = client.patch(
            f"/api/visits/{visit['id']}",
            json={"receptionist": "不存在的接待人昵称"},
            headers=other_headers,
        )
        assert invalid_receptionist.status_code == 400

        set_view_only = client.put("/api/position-permissions/full", json={
            "position": position["name"],
            "pages": ["class-records", "daily-activities"],
            "edit_permissions": {"visits": "view", "activities": "own"},
        })
        assert set_view_only.status_code == 200
        view_only_update = client.patch(
            f"/api/visits/{visit['id']}",
            json={"goal": "仅浏览账号不能修改"},
            headers=other_headers,
        )
        assert view_only_update.status_code == 403
    finally:
        client.delete(f"/api/accounts/{account['id']}")
        client.delete(f"/api/positions/{position['id']}")
        client.delete(f"/api/visits/{visit['id']}")


def test_visit_export_marks_leader_row_yellow(client, created_customer):
    date = "2026-09-05"
    visit_response = client.post("/api/visits", json={
        "visit_date": date,
        "customer_id": created_customer["id"],
        "is_leader": True,
    })
    assert visit_response.status_code == 200, visit_response.text
    visit = visit_response.json()
    try:
        export_response = client.get(f"/api/visits/export?date={date}")
        assert export_response.status_code == 200, export_response.text
        sheet = load_workbook(BytesIO(export_response.content)).active
        leader_row = next(
            row for row in sheet.iter_rows(min_row=2)
            if row[1].value == created_customer["nickname"]
        )
        headers = [cell.value for cell in sheet[1]]
        assert "参与次数" not in headers
        assert "组长获得的信息" not in headers
        assert all(cell.fill.fgColor.rgb == "FFFEFA53" for cell in leader_row)
        nickname_cell = leader_row[headers.index("客户昵称")]
        assert nickname_cell.font.bold is True
        assert nickname_cell.font.size == 13
        leader_status_cell = leader_row[headers.index("组长情况")]
        assert leader_status_cell.font.bold is True
        assert leader_status_cell.font.size == 13
        assert all(cell.alignment.horizontal == "center" for row in sheet.iter_rows() for cell in row)
        goal_column = sheet.cell(1, headers.index("目标") + 1).column_letter
        assert sheet.column_dimensions[goal_column].width == 20
    finally:
        client.delete(f"/api/visits/{visit['id']}")


def test_assigned_course_teacher_can_edit_course_content(client, created_customer):
    suffix = uuid.uuid4().hex[:10]
    date = "2026-09-02"
    record_response = client.post("/api/class-records", json={
        "date": date,
        "course_id": f"teacher-permission-{suffix}",
        "course_name": "老师权限测试",
        "teacher_ids": [created_customer["id"]],
    })
    assert record_response.status_code == 200, record_response.text
    record = record_response.json()

    other_record_response = client.post("/api/class-records", json={
        "date": date,
        "course_id": f"other-teacher-permission-{suffix}",
        "course_name": "非本人授课测试",
    })
    assert other_record_response.status_code == 200, other_record_response.text
    other_record = other_record_response.json()

    position_response = client.post("/api/positions", json={"name": f"老师配置角色_{suffix}"})
    assert position_response.status_code == 200
    position = position_response.json()
    permission_response = client.put("/api/position-permissions/full", json={
        "position": position["name"],
        "pages": ["daily-activities"],
        "edit_permissions": {
            "activities": "view",
            "activity_teachers": "own",
            "activity_participants": "view",
        },
    })
    assert permission_response.status_code == 200

    password = f"pw{suffix}9"
    account_response = client.post("/api/accounts", json={
        "owner": created_customer["nickname"],
        "role": position["name"],
        "username": f"teacher_editor_{suffix}",
        "password": password,
        "enabled": True,
    })
    assert account_response.status_code == 200
    account = account_response.json()
    login = client.post("/api/accounts/login", json={
        "username": account["username"],
        "password": password,
    })
    headers = {"Authorization": f"Bearer {login.json()['token']}"}

    try:
        content_update = client.patch(
            f"/api/class-records/{record['id']}",
            json={"course_name": "授课老师可以修改"},
            headers=headers,
        )
        assert content_update.status_code == 200, content_update.text
        assert content_update.json()["course_name"] == "授课老师可以修改"

        unrelated_update = client.patch(
            f"/api/class-records/{other_record['id']}",
            json={"course_name": "不应允许修改"},
            headers=headers,
        )
        assert unrelated_update.status_code == 403
        unrelated_teacher_assignment = client.patch(
            f"/api/class-records/{other_record['id']}",
            json={"teacher_ids": [created_customer["id"]]},
            headers=headers,
        )
        assert unrelated_teacher_assignment.status_code == 403

        participant_update = client.patch(
            f"/api/class-records/{record['id']}",
            json={"participant_ids": [created_customer["id"]]},
            headers=headers,
        )
        assert participant_update.status_code == 403

        delete_response = client.delete(
            f"/api/class-records/{record['id']}",
            headers=headers,
        )
        assert delete_response.status_code == 403

        disable_teacher_edit = client.put("/api/position-permissions/full", json={
            "position": position["name"],
            "pages": ["daily-activities"],
            "edit_permissions": {
                "activities": "view",
                "activity_teachers": "view",
                "activity_participants": "view",
            },
        })
        assert disable_teacher_edit.status_code == 200
        disabled_update = client.patch(
            f"/api/class-records/{record['id']}",
            json={"course_name": "关闭权限后不能修改"},
            headers=headers,
        )
        assert disabled_update.status_code == 403
    finally:
        client.delete(f"/api/accounts/{account['id']}")
        client.delete(f"/api/positions/{position['id']}")
        client.delete(f"/api/class-records/{record['id']}")
        client.delete(f"/api/class-records/{other_record['id']}")


def test_visit_and_all_schedule_types_use_field_level_creator_permissions(client, created_customer):
    date = "2026-08-20"
    visit = client.post("/api/visits", json={
        "visit_date": date,
        "visit_time": "14:00",
        "customer_id": created_customer["id"],
        "needs": "睡眠调理",
    })
    assert visit.status_code == 200
    assert visit.json()["created_by_id"]
    assert visit.json()["created_by"] == "不闹"

    light_list = client.get(f"/api/visits/light?date={date}")
    assert light_list.status_code == 200
    light_visit = next(item for item in light_list.json() if item["id"] == visit.json()["id"])
    assert light_visit["visit_time"] == "14:00"
    assert light_visit["needs"] == "睡眠调理"
    assert light_visit["arrived_count"] == 0

    records = [
        (
            "/api/class-records",
            {
                "date": date,
                "course_id": "ownership-course",
                "course_name": "归属测试沙龙",
                "participant_ids": [created_customer["id"]],
            },
            {"course_name": "他人不能修改课程名称"},
        ),
        (
            "/api/group-case-sessions",
            {
                "date": date,
                "owner_id": created_customer["id"],
                "owner_name": created_customer["nickname"],
            },
            {"name": "他人不能修改觉醒游戏名称"},
        ),
        (
            "/api/emotional-release-sessions",
            {
                "date": date,
                "owner_id": created_customer["id"],
                "owner_name": created_customer["nickname"],
            },
            {"owner_name": "他人不能修改案主"},
        ),
        (
            "/api/energy-knot-sessions",
            {
                "date": date,
                "owner_id": created_customer["id"],
                "owner_name": created_customer["nickname"],
            },
            {"description": '[{"id":"other","name":"他人","count":3}]'},
        ),
        (
            "/api/internal-course-sessions",
            {
                "date": date,
                "course_type": "归属测试",
                "course_name": "归属测试内部课",
                "participant_ids": [created_customer["id"]],
            },
            {"course_name": "他人不能修改内部课程名称"},
        ),
    ]

    created_records = []
    for path, payload, protected_update in records:
        response = client.post(path, json=payload)
        assert response.status_code == 200, response.text
        body = response.json()
        assert body["created_by_id"]
        assert body["created_by"] == "不闹"
        created_records.append((path, body["id"], protected_update))

    record_type_by_path = {
        "/api/class-records": "class",
        "/api/group-case-sessions": "gcs",
        "/api/emotional-release-sessions": "ers",
        "/api/energy-knot-sessions": "eks",
        "/api/internal-course-sessions": "ics",
    }
    withdrawal_records = [
        (record_type_by_path[path], record_id)
        for path, record_id, _ in created_records
    ]

    account, position, other_headers = _create_other_account_headers(client)
    try:
        visit_id = visit.json()["id"]
        for protected_update in (
            {"visit_date": "2026-08-21"},
            {"visit_time": "15:00"},
            {"customer_id": "other-customer"},
            {"referrer_handler": "其他邀约人"},
        ):
            forbidden_visit_update = client.patch(
                f"/api/visits/{visit_id}",
                json=protected_update,
                headers=other_headers,
            )
            assert forbidden_visit_update.status_code == 403
        cancel_visit = client.patch(
            f"/api/visits/{visit_id}",
            json={"cancelled": True},
            headers=other_headers,
        )
        assert cancel_visit.status_code == 200
        assert cancel_visit.json()["cancelled"] is True
        restore_visit = client.patch(
            f"/api/visits/{visit_id}",
            json={"cancelled": False},
            headers=other_headers,
        )
        assert restore_visit.status_code == 200
        assert restore_visit.json()["cancelled"] is False
        forbidden_visit_delete = client.delete(
            f"/api/visits/{visit_id}", headers=other_headers
        )
        assert forbidden_visit_delete.status_code == 403

        other_shared_update = client.patch(
            f"/api/visits/{visit_id}",
            json={
                "arrived": True,
                "arrival_time": "10:30",
                "is_leader": True,
                "feedback": "他人可补充客户信息",
                "healing_notes": "他人可补充跟进点",
            },
            headers=other_headers,
        )
        assert other_shared_update.status_code == 200
        assert other_shared_update.json()["arrived"] is True
        assert other_shared_update.json()["arrival_time"] == "10:30"
        assert other_shared_update.json()["is_leader"] is True
        assert other_shared_update.json()["feedback"].endswith(
            f"{account['owner']}：他人可补充客户信息"
        )
        assert other_shared_update.json()["healing_notes"].endswith(
            f"{account['owner']}：他人可补充跟进点"
        )
        other_private_need = client.patch(
            f"/api/visits/{visit_id}",
            json={"needs": "其他员工自己的来访需求"},
            headers=other_headers,
        )
        assert other_private_need.status_code == 200
        assert other_private_need.json()["needs"] == "其他员工自己的来访需求"
        assert client.get(f"/api/visits/{visit_id}").json()["needs"] == "睡眠调理"

        mixed_visit_update = client.patch(
            f"/api/visits/{visit_id}",
            json={"arrived": False, "needs": "其他员工更新自己的需求"},
            headers=other_headers,
        )
        assert mixed_visit_update.status_code == 200
        assert mixed_visit_update.json()["needs"] == "其他员工更新自己的需求"

        other_reorder = client.post(
            "/api/visits/reorder",
            json={"ids": [visit_id]},
            headers=other_headers,
        )
        assert other_reorder.status_code == 200

        other_activity_reorder = client.post(
            "/api/activity-orders",
            json={
                "date": date,
                "space_id": "",
                "order": [
                    f"{path.rsplit('/', 1)[-1]}-{record_id}"
                    for path, record_id, _ in created_records
                ],
            },
            headers=other_headers,
        )
        assert other_activity_reorder.status_code == 200

        for path, record_id, protected_update in created_records:
            shared_update = client.patch(
                f"{path}/{record_id}",
                json={"is_published": True},
                headers=other_headers,
            )
            assert shared_update.status_code == 200, shared_update.text
            assert shared_update.json()["is_published"] is True

            for forbidden_payload in (
                protected_update,
                {"start_time": "12:00"},
                {"end_time": "13:00"},
                {"teacher_ids": [created_customer["id"]]},
            ):
                forbidden_update = client.patch(
                    f"{path}/{record_id}",
                    json=forbidden_payload,
                    headers=other_headers,
                )
                assert forbidden_update.status_code == 403, forbidden_update.text
            forbidden_delete = client.delete(
                f"{path}/{record_id}", headers=other_headers
            )
            assert forbidden_delete.status_code == 403, forbidden_delete.text

        welfare_update = client.patch(
            f"/api/class-records/{created_records[0][1]}",
            json={"is_public_welfare": True},
            headers=other_headers,
        )
        assert welfare_update.status_code == 403, welfare_update.text

        class_record_id = created_records[0][1]
        participant_update = client.patch(
            f"/api/class-records/{class_record_id}/participants",
            json={"participant_ids": [created_customer["id"]]},
            headers=other_headers,
        )
        assert participant_update.status_code == 200, participant_update.text
        assert participant_update.json()["participant_ids"] == [created_customer["id"]]

        for record_type, record_id in withdrawal_records:
            forbidden_withdrawal = client.post(
                f"/api/activity-withdrawals/{record_type}/{record_id}",
                json={"customer_id": created_customer["id"]},
                headers=other_headers,
            )
            assert forbidden_withdrawal.status_code == 403, forbidden_withdrawal.text

            owner_withdrawal = client.post(
                f"/api/activity-withdrawals/{record_type}/{record_id}",
                json={"customer_id": created_customer["id"]},
            )
            assert owner_withdrawal.status_code == 200, owner_withdrawal.text
            assert created_customer["id"] in owner_withdrawal.json()["withdrawn_participant_ids"]
            forbidden_restore = client.delete(
                f"/api/activity-withdrawals/{record_type}/{record_id}/{created_customer['id']}",
                headers=other_headers,
            )
            assert forbidden_restore.status_code == 403, forbidden_restore.text
            owner_restore = client.delete(
                f"/api/activity-withdrawals/{record_type}/{record_id}/{created_customer['id']}"
            )
            assert owner_restore.status_code == 200, owner_restore.text
            assert created_customer["id"] not in owner_restore.json()["withdrawn_participant_ids"]

        own_update = client.patch(
            f"/api/visits/{visit_id}", json={"needs": "本人可以修改"}
        )
        assert own_update.status_code == 200
        assert own_update.json()["needs"] == "本人可以修改"

        grant_all = client.put("/api/position-permissions/full", json={
            "position": position["name"],
            "pages": ["class-records", "daily-activities"],
            "edit_permissions": {"visits": "all", "activities": "all"},
        })
        assert grant_all.status_code == 200
        permission_logs = client.get(
            "/api/operation-logs",
            params={"entity_id": position["name"]},
        )
        assert permission_logs.status_code == 200
        permission_log = next(
            item
            for item in permission_logs.json()
            if item["path"].startswith("/api/position-permissions/full")
            and "全部记录" in item["content"]
        )
        assert "邀约编辑范围(仅本人录入→全部记录)" in permission_log["content"]
        assert "课表编辑范围(仅本人录入→全部记录)" in permission_log["content"]
        assert "edit_permissions" not in permission_log["content"]

        for record_type, record_id in withdrawal_records:
            allowed_withdrawal = client.post(
                f"/api/activity-withdrawals/{record_type}/{record_id}",
                json={"customer_id": created_customer["id"]},
                headers=other_headers,
            )
            assert allowed_withdrawal.status_code == 200, allowed_withdrawal.text
            allowed_restore = client.delete(
                f"/api/activity-withdrawals/{record_type}/{record_id}/{created_customer['id']}",
                headers=other_headers,
            )
            assert allowed_restore.status_code == 200, allowed_restore.text

        allowed_visit_update = client.patch(
            f"/api/visits/{visit_id}",
            json={"needs": "全量权限可以修改"},
            headers=other_headers,
        )
        assert allowed_visit_update.status_code == 200
        assert allowed_visit_update.json()["needs"] == "全量权限可以修改"
        assert client.get(f"/api/visits/{visit_id}").json()["needs"] == "本人可以修改"

        for path, record_id, protected_update in created_records:
            allowed_update = client.patch(
                f"{path}/{record_id}",
                json=protected_update,
                headers=other_headers,
            )
            assert allowed_update.status_code == 200, allowed_update.text
    finally:
        for path, record_id, _ in created_records:
            client.delete(f"{path}/{record_id}")
        client.delete(f"/api/visits/{visit.json()['id']}")
        client.delete(f"/api/accounts/{account['id']}")
        client.delete(f"/api/positions/{position['id']}")


def test_super_admin_schedule_permission_does_not_merge_private_visit_needs(client, created_customer):
    date = "2026-08-21"
    visit = client.post("/api/visits", json={
        "visit_date": date,
        "customer_id": created_customer["id"],
        "needs": "原始需求",
    })
    assert visit.status_code == 200

    suffix = uuid.uuid4().hex[:10]
    password = f"pw{suffix}9"
    account_response = client.post("/api/accounts", json={
        "owner": f"超级管理员_{suffix}",
        "role": "超级管理员",
        "username": f"super_edit_{suffix}",
        "password": password,
        "enabled": True,
    })
    assert account_response.status_code == 200
    account = account_response.json()
    login = client.post("/api/accounts/login", json={
        "username": account["username"],
        "password": password,
    })
    assert login.status_code == 200
    assert login.json()["edit_permissions"] == position_edit_permission_service.SUPER_ADMIN_PERMISSIONS
    headers = {"Authorization": f"Bearer {login.json()['token']}"}

    try:
        updated = client.patch(
            f"/api/visits/{visit.json()['id']}",
            json={"needs": "超级管理员修改"},
            headers=headers,
        )
        assert updated.status_code == 200
        assert updated.json()["needs"] == "超级管理员修改"
        assert client.get(f"/api/visits/{visit.json()['id']}").json()["needs"] == "原始需求"
    finally:
        client.delete(f"/api/visits/{visit.json()['id']}")
        client.delete(f"/api/accounts/{account['id']}")
