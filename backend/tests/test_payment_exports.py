import io

from openpyxl import load_workbook


def test_payment_export_contains_all_types_and_sorts_by_deal_date(client, created_customer):
    card_id = ""
    group_case_id = ""
    try:
        card_response = client.post("/api/membership-cards", json={
            "customer_id": created_customer["id"],
            "nickname": created_customer["nickname"],
            "card_type": "次卡",
            "price": 198,
            "effective_date": "2026-08-01",
            "duration_type": "month",
            "duration_value": 1,
            "total_count": 1,
            "remaining_count": 1,
            "deal_date": "2026-08-01",
        })
        assert card_response.status_code == 200
        card_id = card_response.json()["id"]

        group_case_response = client.post("/api/group-cases", json={
            "customer_id": created_customer["id"],
            "nickname": created_customer["nickname"],
            "purchase_count": 2,
            "amount": 799,
            "deal_date": "2026-08-09",
        })
        assert group_case_response.status_code == 200
        group_case_id = group_case_response.json()["id"]

        export_response = client.get("/api/payment-exports/export")
        assert export_response.status_code == 200
        assert export_response.headers["content-type"].startswith(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        workbook = load_workbook(io.BytesIO(export_response.content), read_only=True)
        worksheet = workbook["全部付费记录"]
        rows = list(worksheet.iter_rows(values_only=True))
        assert rows[0][:6] == ("成交日期", "付费类型", "用户昵称", "项目内容", "数量/期限", "金额")

        test_rows = [row for row in rows[1:] if row[2] == created_customer["nickname"]]
        assert [row[0] for row in test_rows[:2]] == ["2026-08-09", "2026-08-01"]
        assert {row[1] for row in test_rows} >= {"会员卡", "觉醒游戏"}

        day_response = client.get("/api/payment-exports/export?range_type=day&period=2026-08-09")
        assert day_response.status_code == 200
        day_rows = list(
            load_workbook(io.BytesIO(day_response.content), read_only=True)["全部付费记录"]
            .iter_rows(values_only=True)
        )
        assert [row[0] for row in day_rows[1:] if row[2] == created_customer["nickname"]] == ["2026-08-09"]

        month_response = client.get("/api/payment-exports/export?range_type=month&period=2026-08")
        assert month_response.status_code == 200
        month_rows = list(
            load_workbook(io.BytesIO(month_response.content), read_only=True)["全部付费记录"]
            .iter_rows(values_only=True)
        )
        assert [row[0] for row in month_rows[1:] if row[2] == created_customer["nickname"]][:2] == [
            "2026-08-09",
            "2026-08-01",
        ]

        year_response = client.get("/api/payment-exports/export?range_type=year&period=2026")
        assert year_response.status_code == 200
        year_rows = list(
            load_workbook(io.BytesIO(year_response.content), read_only=True)["全部付费记录"]
            .iter_rows(values_only=True)
        )
        assert [row[0] for row in year_rows[1:] if row[2] == created_customer["nickname"]][:2] == [
            "2026-08-09",
            "2026-08-01",
        ]

        custom_response = client.get(
            "/api/payment-exports/export?range_type=custom&date_from=2026-08-02&date_to=2026-08-10"
        )
        assert custom_response.status_code == 200
        custom_rows = list(
            load_workbook(io.BytesIO(custom_response.content), read_only=True)["全部付费记录"]
            .iter_rows(values_only=True)
        )
        assert [row[0] for row in custom_rows[1:] if row[2] == created_customer["nickname"]] == ["2026-08-09"]

        empty_response = client.get("/api/payment-exports/export?range_type=year&period=1900")
        assert empty_response.status_code == 404
        assert empty_response.json()["detail"] == "所选时间范围内暂无付费记录"
    finally:
        if group_case_id:
            client.delete(f"/api/group-cases/{group_case_id}")
        if card_id:
            client.delete(f"/api/membership-cards/{card_id}")
